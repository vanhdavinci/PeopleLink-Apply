"""Master File: map hồ sơ nguồn → Excel 2 sheet (DANH SÁCH + LỊCH LÀM VIỆC)."""
from __future__ import annotations

import io
import re
from datetime import date, datetime
from itertools import groupby
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from app.config import TEMPLATES_DIR

TEMPLATE_PATH = TEMPLATES_DIR / "master_file_llv_mtf.xlsx"
TRAINING_TEMPLATE_PATH = TEMPLATES_DIR / "master_file_list_training.xlsx"

SHEET_DANH_SACH = "DANH SÁCH"
SHEET_LICH = "LỊCH LÀM VIỆC"
SHEET_TRAINING_LIST = "Training List"
SHEET_DETAIL = "Detail"

_THIN_BORDER = Border(
    left=Side(style="thin", color="000000"),
    right=Side(style="thin", color="000000"),
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000"),
)

# Cột sheet DANH SÁCH (đúng thứ tự template)
DANH_SACH_COLUMNS: list[str] = [
    "Họ tên",
    "Giới tính",
    "Tình trạng hôn nhân",
    "Ngày sinh",
    "Nơi sinh",
    "Nguyên quán",
    "CMND/CCCD",
    "Ngày cấp",
    "Nơi cấp",
    "Di động",
    "E-mail",
    "SĐT người thân",
    "Địa chỉ tạm trú",
    "Địa chỉ thường trú",
    "Dân tộc",
    "Quốc tịch",
    "Trình độ học vấn",
    "Chiều cao",
    "Cân nặng",
    "Size áo",
    "Size quần",
    "Size giầy",
    "MST",
    "Ngân hàng",
    "Chi nhánh",
    "Số TK",
    "Tên TK",
    "Vị trí - Trình độ",
    "Người phụ trách",
    "Khu vực / Siêu thị đăng ký làm việc",
    "Ngày Onboard",
    "Note",
]

# Nguồn → cột DANH SÁCH (cùng tên trừ vài alias)
_SOURCE_ALIASES: dict[str, tuple[str, ...]] = {
    "Họ tên": ("Họ tên", "Ho ten", "FullName"),
    "CMND/CCCD": ("CMND/CCCD", "CMND/CCCD cũ", "CCCD", "CMND"),
    "E-mail": ("E-mail", "Email", "e-mail"),
    "Note": ("Note", "Ghi chú", "Notes"),
}

LICH_COLUMNS: list[str] = [
    "STT",
    "Address",
    "Detail",
    "Day",
    "Time",
    "SUP/ PG ĐĂNG KÝ",
    "SĐT",
    "STATUS",
]

# Cột map sang sheet Training List (header row 3 trong template)
TRAINING_LIST_COLUMNS: list[str] = [
    "STT",
    "Họ tên",
    "SDT",
    "Vị trí",
    "Khu vực",
    "Siêu thị/ Địa điểm",
]


def _norm_header(text: Any) -> str:
    s = str(text or "").strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def _cell_str(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")
    text = str(value).strip()
    if text.lower() in {"nan", "none", "null", "null@null"}:
        return ""
    return text


def _find_sheet(wb, *candidates: str) -> Worksheet:
    names = {str(n).strip().casefold(): n for n in wb.sheetnames}
    for cand in candidates:
        key = cand.strip().casefold()
        if key in names:
            return wb[names[key]]
        # partial: startswith
        for k, real in names.items():
            if k.startswith(key) or key.startswith(k):
                return wb[real]
    # fallback first sheet
    return wb[wb.sheetnames[0]]


def _detect_header_row(ws: Worksheet, *, scan_rows: int = 15) -> int:
    """Tìm dòng header chứa 'Họ tên' hoặc 'CMND'."""
    markers = {_norm_header("Họ tên"), _norm_header("CMND/CCCD"), _norm_header("Di động")}
    for r in range(1, min(scan_rows, ws.max_row or 1) + 1):
        vals = {
            _norm_header(ws.cell(r, c).value)
            for c in range(1, min(80, (ws.max_column or 1) + 1))
        }
        vals.discard("")
        if len(vals & markers) >= 2:
            return r
    return 1


def _read_headers(ws: Worksheet, header_row: int) -> dict[str, int]:
    """normalized header → 1-based col index (first wins)."""
    mapping: dict[str, int] = {}
    for c in range(1, (ws.max_column or 1) + 1):
        raw = ws.cell(header_row, c).value
        if raw is None or str(raw).strip() == "":
            continue
        key = _norm_header(raw)
        if key not in mapping:
            mapping[key] = c
    return mapping


def _lookup_col(headers: dict[str, int], target: str) -> int | None:
    aliases = _SOURCE_ALIASES.get(target, (target,))
    for alias in aliases:
        idx = headers.get(_norm_header(alias))
        if idx:
            return idx
    # exact norm of target
    return headers.get(_norm_header(target))


def _compose_full_name(row_vals: dict[str, Any], headers: dict[str, int], row: int, ws: Worksheet) -> str:
    full = _cell_str(row_vals.get("Họ tên"))
    if full:
        return full
    parts = []
    for key in ("Họ", "Tên đệm", "Tên"):
        col = headers.get(_norm_header(key))
        if col:
            parts.append(_cell_str(ws.cell(row, col).value))
    return " ".join(p for p in parts if p).strip()


def parse_source_workbook(data: bytes | Path) -> list[dict[str, Any]]:
    """Đọc file hồ sơ nguồn → list dict theo cột DANH SÁCH."""
    if isinstance(data, Path):
        wb = load_workbook(data, data_only=True)
    else:
        wb = load_workbook(io.BytesIO(data), data_only=True)

    ws = _find_sheet(wb, "Danh sách", "DANH SÁCH", "Sheet1")
    header_row = _detect_header_row(ws)
    headers = _read_headers(ws, header_row)

    rows: list[dict[str, Any]] = []
    for r in range(header_row + 1, (ws.max_row or header_row) + 1):
        # skip empty rows (no name / no phone / no cccd)
        probe_cols = [
            _lookup_col(headers, "Họ tên"),
            _lookup_col(headers, "Di động"),
            _lookup_col(headers, "CMND/CCCD"),
            headers.get(_norm_header("Họ")),
        ]
        has_any = False
        for col in probe_cols:
            if col and _cell_str(ws.cell(r, col).value):
                has_any = True
                break
        if not has_any:
            continue

        item: dict[str, Any] = {}
        for dest in DANH_SACH_COLUMNS:
            if dest == "Họ tên":
                # gather raw first for compose
                col = _lookup_col(headers, "Họ tên")
                raw_map = {"Họ tên": ws.cell(r, col).value if col else None}
                item[dest] = _compose_full_name(raw_map, headers, r, ws)
                continue
            if dest == "Note":
                col = _lookup_col(headers, "Note")
                item[dest] = _cell_str(ws.cell(r, col).value) if col else ""
                continue
            col = _lookup_col(headers, dest)
            item[dest] = _cell_str(ws.cell(r, col).value) if col else ""
        rows.append(item)

    wb.close()
    return rows


def build_lich_lam_viec(danh_sach_rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Sinh sheet LỊCH LÀM VIỆC cơ bản từ DANH SÁCH (v1 — sẽ bổ sung sau)."""
    out: list[dict[str, Any]] = []
    for i, row in enumerate(danh_sach_rows, start=1):
        khu_vuc = _cell_str(row.get("Khu vực / Siêu thị đăng ký làm việc"))
        out.append(
            {
                "STT": i,
                "Address": "",
                "Detail": khu_vuc,
                "Day": "",
                "Time": "",
                "SUP/ PG ĐĂNG KÝ": _cell_str(row.get("Họ tên")),
                "SĐT": _cell_str(row.get("Di động")),
                "STATUS": "",
            }
        )
    return out


def format_schedule_days(dates: list[date | datetime]) -> str:
    """Format như template: `25,26/07/2026` (gom theo tháng)."""
    cleaned: list[date] = []
    for raw in dates or []:
        if isinstance(raw, datetime):
            cleaned.append(raw.date())
        elif isinstance(raw, date):
            cleaned.append(raw)
    if not cleaned:
        return ""
    uniq = sorted(set(cleaned))
    parts: list[str] = []
    for (year, month), group in groupby(uniq, key=lambda d: (d.year, d.month)):
        days = [d.day for d in group]
        if len(days) == 1:
            parts.append(f"{days[0]:02d}/{month:02d}/{year}")
        else:
            day_part = ",".join(str(d) for d in days)
            parts.append(f"{day_part}/{month:02d}/{year}")
    return ", ".join(parts)


def format_schedule_times(slots: list[str]) -> str:
    """Format như template: `6h-11h, 16h-20h`."""
    cleaned = []
    for slot in slots or []:
        text = re.sub(r"\s+", "", str(slot or "").strip().lower())
        if not text:
            continue
        # normalize 06h-11h → 6h-11h
        text = re.sub(r"\b0+(\d)", r"\1", text)
        if text not in cleaned:
            cleaned.append(text)
    return ", ".join(cleaned)


def make_time_slot(start_hour: int, end_hour: int) -> str:
    start = max(0, min(23, int(start_hour)))
    end = max(0, min(24, int(end_hour)))
    return f"{start}h-{end}h"


def apply_lich_bulk_fields(
    lich_rows: list[dict[str, Any]],
    *,
    address: str | None = None,
    day: str | None = None,
    time: str | None = None,
) -> list[dict[str, Any]]:
    """Ghi Address / Day / Time cho toàn bộ dòng lịch."""
    out: list[dict[str, Any]] = []
    for row in lich_rows:
        item = dict(row)
        if address is not None:
            item["Address"] = address
        if day is not None:
            item["Day"] = day
        if time is not None:
            item["Time"] = time
        out.append(item)
    return out


def refresh_master_export(result: dict[str, Any]) -> dict[str, Any]:
    """Cập nhật lại xlsx_bytes (+ training) từ danh_sach + lich hiện tại."""
    danh_sach = list(result.get("danh_sach") or [])
    lich = list(result.get("lich") or [])
    payload = export_master_file_bytes(danh_sach, lich_rows=lich)
    training_rows = build_training_list_rows(danh_sach, lich)
    training_bytes = export_training_list_bytes(
        danh_sach_rows=danh_sach,
        lich_rows=lich,
    )
    return {
        **result,
        "danh_sach": danh_sach,
        "lich": lich,
        "training_list": training_rows,
        "row_count": len(danh_sach),
        "lich_count": len(lich),
        "xlsx_bytes": payload,
        "training_xlsx_bytes": training_bytes,
    }


def _short_position(raw: str) -> str:
    """Vị trí - Trình độ → SUP / PG (theo mẫu Training List)."""
    text = _cell_str(raw)
    if not text:
        return ""
    upper = text.upper()
    if re.search(r"\bSUP\b", upper) or upper.startswith("SUP"):
        return "SUP"
    if re.search(r"\bPG\b", upper) or upper.startswith("PG"):
        return "PG"
    if " - " in text:
        return text.split(" - ", 1)[0].strip()
    return text


def build_training_list_rows(
    danh_sach_rows: list[dict[str, Any]],
    lich_rows: list[dict[str, Any]] | None = None,
) -> list[dict[str, Any]]:
    """Map DANH SÁCH + LỊCH → dòng Training List."""
    lich = lich_rows if lich_rows is not None else build_lich_lam_viec(danh_sach_rows)
    out: list[dict[str, Any]] = []
    for i, ds in enumerate(danh_sach_rows):
        ll = lich[i] if i < len(lich) else {}
        out.append(
            {
                "STT": i + 1,
                "Họ tên": _cell_str(ds.get("Họ tên"))
                or _cell_str(ll.get("SUP/ PG ĐĂNG KÝ")),
                "SDT": _cell_str(ds.get("Di động")) or _cell_str(ll.get("SĐT")),
                "Vị trí": _short_position(str(ds.get("Vị trí - Trình độ") or "")),
                "Khu vực": _cell_str(ll.get("Address")),
                "Siêu thị/ Địa điểm": _cell_str(ll.get("Detail"))
                or _cell_str(ds.get("Khu vực / Siêu thị đăng ký làm việc")),
            }
        )
    return out



# Training List: cột dữ liệu map (A-F). Toàn bảng style tới cột P (16) = Đánh Giá.
_TRAINING_HEADER_ROW = 3
_TRAINING_DATA_START = 4
_TRAINING_LAST_COL = 16  # Đánh Giá
_TRAINING_DATA_ROW_HEIGHT = 37.05
_TRAINING_COL_MAP = {
    1: "STT",
    2: "Họ tên",
    3: "SDT",
    4: "Vị trí",
    5: "Khu vực",
    6: "Siêu thị/ Địa điểm",
}


def _snapshot_row_styles(ws: Worksheet, row: int, max_col: int) -> list[dict[str, Any]]:
    """Lưu style 1 dòng mẫu để copy sang dòng mới."""
    from copy import copy

    styles: list[dict[str, Any]] = []
    for c in range(1, max_col + 1):
        cell = ws.cell(row, c)
        styles.append(
            {
                "font": copy(cell.font)
                if cell.has_style
                else Font(name="Times New Roman", size=11),
                "border": copy(cell.border) if cell.has_style else _THIN_BORDER,
                "fill": copy(cell.fill)
                if cell.has_style
                else PatternFill(fill_type=None),
                "alignment": copy(cell.alignment)
                if cell.has_style
                else Alignment(horizontal="center", vertical="center"),
                "number_format": cell.number_format,
            }
        )
    return styles


def _apply_style_dict(cell: Any, style: dict[str, Any]) -> None:
    cell.font = style["font"]
    cell.border = style["border"]
    cell.fill = style["fill"]
    cell.alignment = style["alignment"]
    cell.number_format = style["number_format"]


def _fill_training_list_sheet(
    ws: Worksheet,
    training_rows: list[dict[str, Any]],
    *,
    title: str | None = None,
) -> None:
    """
    Ghi data vào Training List, giữ nguyên format template
    (font/size/width/header màu xanh / merge / border all).
    """
    header_row = _TRAINING_HEADER_ROW
    last_col = _TRAINING_LAST_COL

    if title:
        title_cell = ws.cell(1, 1)
        title_cell.value = title
        title_cell.font = Font(name="Times New Roman", size=12, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")

    sample_row = _TRAINING_DATA_START
    if (ws.max_row or 0) < sample_row:
        sample_row = header_row
    styles = _snapshot_row_styles(ws, sample_row, last_col)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(last_col):
        styles[c]["border"] = _THIN_BORDER
        styles[c]["font"] = Font(name="Times New Roman", size=11)
        styles[c]["alignment"] = center
        styles[c]["fill"] = PatternFill(fill_type="solid", fgColor="FFFFFF")

    # Xóa data cũ, không đụng header / column width / merge header
    if ws.max_row and ws.max_row > header_row:
        ws.delete_rows(header_row + 1, ws.max_row - header_row)

    # Header: chỉ bổ sung Border All, giữ nguyên màu nền/chữ của template
    for c in range(1, last_col + 1):
        ws.cell(header_row, c).border = _THIN_BORDER
    if ws.row_dimensions[header_row].height is None:
        ws.row_dimensions[header_row].height = 45.0

    for i, row in enumerate(training_rows):
        r = _TRAINING_DATA_START + i
        ws.row_dimensions[r].height = _TRAINING_DATA_ROW_HEIGHT
        for c in range(1, last_col + 1):
            cell = ws.cell(r, c)
            key = _TRAINING_COL_MAP.get(c)
            cell.value = row.get(key, "") if key else None
            _apply_style_dict(cell, styles[c - 1])
            if c in (2, 6) and key:
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )

    last_data_row = header_row + len(training_rows)
    try:
        ws.auto_filter.ref = f"A{header_row}:P{max(last_data_row, header_row)}"
    except Exception:
        pass


def export_training_list_bytes(
    *,
    danh_sach_rows: list[dict[str, Any]],
    lich_rows: list[dict[str, Any]] | None = None,
    template_path: Path | None = None,
    title: str | None = None,
) -> bytes:
    """
    Xuất file List training:
    - Sheet Detail: giữ nguyên
    - Sheet Training List: map dữ liệu (giữ format mẫu)
    - Các sheet khác: bỏ (chỉ còn 2 sheet)
    """
    training_rows = build_training_list_rows(danh_sach_rows, lich_rows)
    path = template_path or TRAINING_TEMPLATE_PATH

    if path.is_file():
        wb = load_workbook(path)
        detail_name = _resolve_sheet_name(wb, SHEET_DETAIL)
        train_name = _resolve_sheet_name(wb, SHEET_TRAINING_LIST)
        keep: set[str] = set()
        if detail_name in wb.sheetnames:
            keep.add(detail_name)
        else:
            wb.create_sheet(SHEET_DETAIL)
            detail_name = SHEET_DETAIL
            keep.add(detail_name)
        if train_name in wb.sheetnames:
            keep.add(train_name)
        else:
            wb.create_sheet(SHEET_TRAINING_LIST, 0)
            train_name = SHEET_TRAINING_LIST
            keep.add(train_name)
        for name in list(wb.sheetnames):
            if name not in keep:
                del wb[name]

        _fill_training_list_sheet(wb[train_name], training_rows, title=title)
        try:
            idx = wb.sheetnames.index(train_name)
            if idx > 0:
                wb.move_sheet(train_name, offset=-idx)
        except Exception:
            pass
    else:
        wb = Workbook()
        ws_tl = wb.active
        ws_tl.title = SHEET_TRAINING_LIST
        headers = [
            "STT",
            "Họ tên",
            "SDT",
            "Vị trí",
            "Khu vực",
            "Siêu thị/ Địa điểm",
            "Điểm Danh",
            None,
            "Đúng Giờ\n(1đ)",
            "Thái Độ Học (1đ)",
            "Phát Biểu (1đ)",
            "Tham Dự Đủ\n(1đ)",
            "Ngoại Hình\n(1đ)",
            "Bài Test (5đ)",
            "Tổng Điểm (10đ)",
            "Đánh Giá",
        ]
        for c, name in enumerate(headers, start=1):
            cell = ws_tl.cell(3, c)
            cell.value = name
            cell.fill = PatternFill(fill_type="solid", fgColor="375623")
            cell.font = Font(
                name="Times New Roman", size=11, bold=True, color="FFFFFF"
            )
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = _THIN_BORDER
        _fill_training_list_sheet(ws_tl, training_rows, title=title)
        wb.create_sheet(SHEET_DETAIL)

    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def _clear_data_rows(ws: Worksheet, *, header_row: int = 1) -> None:
    # Bỏ Excel Table (hay để lại dòng trống tô màu / banded rows).
    try:
        for table_name in list(ws.tables.keys()):
            del ws.tables[table_name]
    except Exception:
        pass
    if ws.max_row and ws.max_row > header_row:
        ws.delete_rows(header_row + 1, ws.max_row - header_row)
    # Xóa style dòng còn sót
    for key in list(ws.row_dimensions.keys()):
        if key > header_row:
            try:
                del ws.row_dimensions[key]
            except Exception:
                pass


def _is_yellow_fill(fill: Any) -> bool:
    if fill is None or not getattr(fill, "patternType", None):
        return False
    color = getattr(fill, "fgColor", None)
    if color is None:
        return False
    rgb = getattr(color, "rgb", None)
    if rgb is not None:
        text = str(rgb).upper().replace("#", "")
        if text.endswith("FFFF00"):
            return True
    # indexed yellow phổ biến trong Excel
    indexed = getattr(color, "indexed", None)
    if indexed in (5, 6, 13, 43, 44):  # yellow-ish indexes
        return True
    return False


def _strip_yellow_fills(ws: Worksheet, *, header_row: int = 1) -> None:
    """Bỏ nền vàng còn sót từ template (giữ style header)."""
    max_r = max(ws.max_row or header_row, header_row + 50)
    max_c = max(ws.max_column or 1, 40)
    clear = PatternFill(fill_type=None)
    for r in range(header_row + 1, max_r + 1):
        for c in range(1, max_c + 1):
            cell = ws.cell(r, c)
            if _is_yellow_fill(cell.fill):
                cell.fill = clear


def _apply_all_borders(
    ws: Worksheet,
    *,
    header_row: int,
    n_cols: int,
    n_data_rows: int,
) -> None:
    """All Borders cho header + toàn bộ dòng dữ liệu."""
    last_row = header_row + max(0, n_data_rows)
    for r in range(header_row, last_row + 1):
        for c in range(1, n_cols + 1):
            ws.cell(r, c).border = _THIN_BORDER


def _apply_times_new_roman(
    ws: Worksheet,
    *,
    header_row: int,
    n_cols: int,
    n_data_rows: int,
    header_font_color: str | None = "FFFFFF",
) -> None:
    """Toàn bộ chữ trong bảng = Times New Roman."""
    last_row = header_row + max(0, n_data_rows)
    for c in range(1, n_cols + 1):
        header_cell = ws.cell(header_row, c)
        size = header_cell.font.size if header_cell.font and header_cell.font.size else 11
        bold = True if header_cell.font is None else bool(header_cell.font.bold)
        if header_font_color:
            header_cell.font = Font(
                name="Times New Roman",
                bold=True,
                size=size,
                color=header_font_color,
            )
        else:
            header_cell.font = Font(name="Times New Roman", bold=bold or True, size=size)
    for r in range(header_row + 1, last_row + 1):
        for c in range(1, n_cols + 1):
            cell = ws.cell(r, c)
            size = cell.font.size if cell.font and cell.font.size else 11
            cell.font = Font(name="Times New Roman", size=size)


def _write_rows(
    ws: Worksheet,
    columns: list[str],
    rows: list[dict[str, Any]],
    *,
    header_row: int = 1,
    center_columns: tuple[str, ...] = (),
    header_font_color: str | None = "FFFFFF",
    force_white_data_fill: bool = True,
) -> None:
    white = PatternFill(fill_type="solid", fgColor="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    center_idx = {
        i for i, name in enumerate(columns, start=1) if name in center_columns
    }
    # ensure headers
    for c, name in enumerate(columns, start=1):
        cell = ws.cell(header_row, c)
        if not cell.value:
            cell.value = name
        if c in center_idx:
            cell.alignment = center
    for i, row in enumerate(rows):
        r = header_row + 1 + i
        for c, name in enumerate(columns, start=1):
            cell = ws.cell(r, c)
            cell.value = row.get(name, "")
            if force_white_data_fill:
                cell.fill = white
            if c in center_idx:
                cell.alignment = center
    # Xóa dòng thừa (tránh ô trống còn style vàng)
    last = header_row + len(rows)
    if ws.max_row and ws.max_row > last:
        ws.delete_rows(last + 1, ws.max_row - last)
    _strip_yellow_fills(ws, header_row=header_row)
    _apply_all_borders(
        ws,
        header_row=header_row,
        n_cols=len(columns),
        n_data_rows=len(rows),
    )
    _apply_times_new_roman(
        ws,
        header_row=header_row,
        n_cols=len(columns),
        n_data_rows=len(rows),
        header_font_color=header_font_color,
    )


def _resolve_sheet_name(wb, preferred: str) -> str:
    for name in wb.sheetnames:
        if name.strip().casefold() == preferred.strip().casefold():
            return name
    # create if missing
    return preferred


def export_master_file_bytes(
    danh_sach_rows: list[dict[str, Any]],
    *,
    lich_rows: list[dict[str, Any]] | None = None,
    template_path: Path | None = None,
) -> bytes:
    """Xuất Excel 2 sheet theo template (giữ style nếu có file mẫu)."""
    lich = lich_rows if lich_rows is not None else build_lich_lam_viec(danh_sach_rows)
    path = template_path or TEMPLATE_PATH

    if path.is_file():
        wb = load_workbook(path)
        ds_name = _resolve_sheet_name(wb, SHEET_DANH_SACH)
        ll_name = _resolve_sheet_name(wb, SHEET_LICH)
        if ds_name not in wb.sheetnames:
            wb.create_sheet(SHEET_DANH_SACH, 0)
            ds_name = SHEET_DANH_SACH
        if ll_name not in wb.sheetnames:
            wb.create_sheet(SHEET_LICH)
            ll_name = SHEET_LICH
        ws_ds = wb[ds_name]
        ws_ll = wb[ll_name]
        _clear_data_rows(ws_ds, header_row=1)
        _clear_data_rows(ws_ll, header_row=1)
        _write_rows(ws_ds, DANH_SACH_COLUMNS, danh_sach_rows, header_row=1)
        _write_rows(
            ws_ll,
            LICH_COLUMNS,
            lich,
            header_row=1,
            center_columns=("STT",),
        )
    else:
        wb = Workbook()
        ws_ds = wb.active
        ws_ds.title = SHEET_DANH_SACH
        ws_ll = wb.create_sheet(SHEET_LICH)
        _write_rows(ws_ds, DANH_SACH_COLUMNS, danh_sach_rows, header_row=1)
        _write_rows(
            ws_ll,
            LICH_COLUMNS,
            lich,
            header_row=1,
            center_columns=("STT",),
        )
    buf = io.BytesIO()
    wb.save(buf)
    wb.close()
    return buf.getvalue()


def process_master_upload(file_bytes: bytes) -> dict[str, Any]:
    """Parse upload → map → build export bytes (LLV&MTF + List training)."""
    rows = parse_source_workbook(file_bytes)
    lich = build_lich_lam_viec(rows)
    payload = export_master_file_bytes(rows, lich_rows=lich)
    training_rows = build_training_list_rows(rows, lich)
    training_bytes = export_training_list_bytes(
        danh_sach_rows=rows,
        lich_rows=lich,
    )
    return {
        "row_count": len(rows),
        "lich_count": len(lich),
        "danh_sach": rows,
        "lich": lich,
        "training_list": training_rows,
        "xlsx_bytes": payload,
        "training_xlsx_bytes": training_bytes,
    }
