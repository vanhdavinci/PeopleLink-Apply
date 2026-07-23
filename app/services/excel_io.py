"""Excel import/export for candidate apply rows."""
from __future__ import annotations

import json
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from app.config import TEMPLATES_DIR
from app.db import get_conn
from app.services.address_resolve import enrich_row_address
from app.services.locations import resolve_location_fields

# Columns user fills in Excel
CANDIDATE_TEMPLATE_COLUMNS: list[str] = [
    "ApplyURL",
    "FullName",
    "Mobile",
    "Sex",
    "Birthday",
    "AcademicLevel",
    "Email",
    "FullAddress",
    "Height",
    "Weight",
    "ApplyExperienceNote",
    "WishWorkplace",
    "ProjectHeadcountType",
]

# Extra fields stored in payload / shown on grid after resolve
ADDRESS_RUNTIME_COLUMNS: list[str] = [
    "AddrTmpStreet",
    "AddrTmpProvince",
    "AddrTmpDistrict",
    "AddrTmpWard",
    "AddressStatus",
    "AddressIcon",
    "AddressNote",
]

PAYLOAD_COLUMNS: list[str] = list(
    dict.fromkeys([*CANDIDATE_TEMPLATE_COLUMNS, *ADDRESS_RUNTIME_COLUMNS])
)

COLUMN_GUIDE: dict[str, str] = {
    "ApplyURL": (
        "Link ứng tuyển → lấy HeadcountReqRecruiterID, ví dụ "
        "https://recruit.peoplelinkvietnam.com/ProjectHeadcount/ApplyRequest/2239"
    ),
    "FullName": "Họ tên ứng viên",
    "Mobile": "Số điện thoại, ví dụ 0934 213 321",
    "Sex": "1 = Nam, 2 = Nữ",
    "Birthday": "Ngày sinh DD/MM/YYYY, ví dụ 07/08/2000",
    "AcademicLevel": (
        "Trình độ: Tiểu học | Trung học cơ sở | Trung học phổ thông | 12/12 | "
        "Cao đẳng | Đại học | Thạc sĩ | Tiến sĩ | Giáo sư | 9/12 | 10/12 | 11/12 | Trung cấp"
    ),
    "Email": "Email ứng viên",
    "FullAddress": (
        "Một dòng địa chỉ thuần. Cũ: '123 Nguyễn Huệ, Phường Bến Nghé, Quận 1, "
        "Thành phố Hồ Chí Minh'. Mới (không huyện): 'Phường Ngọc Hà, Thành phố Hà Nội'. "
        "App sẽ tách; địa chỉ mới đánh dấu ★ để chọn gợi ý map sang cũ."
    ),
    "Height": "Chiều cao (cm)",
    "Weight": "Cân nặng (kg)",
    "ApplyExperienceNote": "Kinh nghiệm làm việc",
    "WishWorkplace": "Khu vực / siêu thị đăng ký (có thể để trống)",
    "ProjectHeadcountType": "Loại project headcount, thường = 3 (Adhoc). Để trống → 3",
}

EXAMPLE_ROW: dict[str, str] = {
    "ApplyURL": "https://recruit.peoplelinkvietnam.com/ProjectHeadcount/ApplyRequest/2239",
    "FullName": "Tuấn Trần",
    "Mobile": "0934 213 321",
    "Sex": "1",
    "Birthday": "07/08/2000",
    "AcademicLevel": "Đại học",
    "Email": "kimnganheo@gmail.com",
    "FullAddress": "123, Phường Châu Phú B, Thành phố Châu Đốc, Tỉnh An Giang",
    "Height": "170",
    "Weight": "59",
    "ApplyExperienceNote": "không co",
    "WishWorkplace": "",
    "ProjectHeadcountType": "3",
}

ACADEMIC_LEVELS: list[str] = [
    "Tiểu học",
    "Trung học cơ sở",
    "Trung học phổ thông",
    "12/12",
    "Cao đẳng",
    "Đại học",
    "Thạc sĩ",
    "Tiến sĩ",
    "Giáo sư",
    "9/12",
    "10/12",
    "11/12",
    "Trung cấp",
]

DEFAULT_TEMPLATE_PATH = TEMPLATES_DIR / "candidates_import_template.xlsx"


def _now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _style_header(ws, row: int = 1) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(color="FFFFFF", bold=True)
    for col in range(1, len(CANDIDATE_TEMPLATE_COLUMNS) + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(wrap_text=True, vertical="center")


def build_candidate_template_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidates"

    for idx, col in enumerate(CANDIDATE_TEMPLATE_COLUMNS, start=1):
        ws.cell(row=1, column=idx, value=col)
    _style_header(ws, 1)

    for idx, col in enumerate(CANDIDATE_TEMPLATE_COLUMNS, start=1):
        ws.cell(row=2, column=idx, value=EXAMPLE_ROW.get(col, ""))

    sex_dv = DataValidation(
        type="list",
        formula1='"1,2"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Sex",
        error="Chỉ nhập 1 (Nam) hoặc 2 (Nữ)",
    )
    sex_dv.add("D3:D1000")
    ws.add_data_validation(sex_dv)

    academic_options = '"' + ",".join(ACADEMIC_LEVELS) + '"'
    academic_dv = DataValidation(
        type="list",
        formula1=academic_options,
        allow_blank=True,
    )
    academic_dv.add("F3:F1000")
    ws.add_data_validation(academic_dv)

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"
    for idx, col in enumerate(CANDIDATE_TEMPLATE_COLUMNS, start=1):
        width = max(14, min(42, len(col) + 8))
        if col in (
            "ApplyURL",
            "FullAddress",
            "Email",
        ):
            width = 42
        ws.column_dimensions[get_column_letter(idx)].width = width

    guide = wb.create_sheet("HuongDan")
    guide["A1"] = "Cột"
    guide["B1"] = "Hướng dẫn điền"
    guide["A1"].font = Font(bold=True)
    guide["B1"].font = Font(bold=True)
    for i, col in enumerate(CANDIDATE_TEMPLATE_COLUMNS, start=2):
        guide.cell(row=i, column=1, value=col)
        guide.cell(row=i, column=2, value=COLUMN_GUIDE.get(col, ""))
    guide.column_dimensions["A"].width = 22
    guide.column_dimensions["B"].width = 90

    notes = wb.create_sheet("LuuY")
    notes["A1"] = "Lưu ý khi điền template"
    notes["A1"].font = Font(bold=True, size=14)
    notes_lines = [
        "1. Chỉ sửa sheet Candidates. Giữ nguyên tên cột hàng 1.",
        "2. Hàng 2 là ví dụ — có thể xóa hoặc ghi đè bằng dữ liệu thật.",
        "3. ApplyURL: mỗi dòng có thể cùng 1 link hoặc khác link tùy vị trí tuyển.",
        "4. RecruiterPIC / HeadcountRequestID KHÔNG có trong Excel — lấy từ User trong app.",
        "5. FullAddress: một dòng địa chỉ thuần. App tự tách Province/District/Ward.",
        "   - Địa chỉ cũ (có huyện/quận) → tự tách.",
        "   - Địa chỉ mới (không huyện) → đánh dấu ★, chọn gợi ý rồi xác nhận map sang cũ.",
        "6. Birthday dùng DD/MM/YYYY.",
        "7. Sex: 1 = Nam, 2 = Nữ.",
        "8. Import file trong app → review lưới → xử lý ★ nếu có → Lưu / Đẩy.",
    ]
    for i, line in enumerate(notes_lines, start=3):
        notes.cell(row=i, column=1, value=line)
    notes.column_dimensions["A"].width = 100
    return wb


def export_candidate_template_bytes() -> bytes:
    wb = build_candidate_template_workbook()
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_candidate_template_file(path: Path | None = None) -> Path:
    out = Path(path) if path else DEFAULT_TEMPLATE_PATH
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = build_candidate_template_workbook()
    wb.save(out)
    return out


def candidate_template_columns() -> list[str]:
    return list(CANDIDATE_TEMPLATE_COLUMNS)


def _cell_str(value: Any) -> str:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y")
    text = str(value).strip()
    if text.lower() in {"nan", "none", "nat"}:
        return ""
    if isinstance(value, float) and text.endswith(".0"):
        text = text[:-2]
    return text


def parse_candidates_excel(file_bytes: bytes | BytesIO) -> list[dict[str, str]]:
    source: BytesIO
    if isinstance(file_bytes, bytes):
        source = BytesIO(file_bytes)
    else:
        source = file_bytes
    df = pd.read_excel(source, sheet_name="Candidates", dtype=object)
    df.columns = [str(c).strip() for c in df.columns]

    # New template uses FullAddress; still accept legacy Prov/Dist/Ward files.
    required = [c for c in CANDIDATE_TEMPLATE_COLUMNS if c != "FullAddress"]
    missing = [c for c in required if c not in df.columns]
    has_full = "FullAddress" in df.columns
    has_legacy = all(
        c in df.columns for c in ("AddrTmpProvince", "AddrTmpDistrict", "AddrTmpWard")
    )
    if missing:
        raise ValueError(
            "File thiếu cột: " + ", ".join(missing) + ". Dùng đúng Excel template."
        )
    if not has_full and not has_legacy:
        raise ValueError(
            "File cần cột FullAddress (mới) hoặc AddrTmpProvince/District/Ward (cũ)."
        )

    rows: list[dict[str, str]] = []
    for _, series in df.iterrows():
        row = {col: "" for col in PAYLOAD_COLUMNS}
        for col in CANDIDATE_TEMPLATE_COLUMNS:
            if col in df.columns:
                row[col] = _cell_str(series.get(col))
        for col in ADDRESS_RUNTIME_COLUMNS:
            if col in df.columns:
                row[col] = _cell_str(series.get(col))
        if not any(row.values()):
            continue
        if not row.get("FullName") and not row.get("Mobile"):
            continue
        if not row.get("ProjectHeadcountType"):
            row["ProjectHeadcountType"] = "3"
        # Build FullAddress from legacy columns if needed
        if not row.get("FullAddress") and has_legacy:
            bits = [
                _cell_str(series.get("AddrTmpStreet")),
                _cell_str(series.get("AddrTmpWard")),
                _cell_str(series.get("AddrTmpDistrict")),
                _cell_str(series.get("AddrTmpProvince")),
            ]
            # strip CODE| from legacy inserted values for display line
            cleaned = []
            for b in bits:
                if not b:
                    continue
                cleaned.append(b.split("|", 1)[-1].strip() if "|" in b else b)
            row["FullAddress"] = ", ".join(cleaned)
        row = enrich_row_address(row, force=True)
        rows.append(row)
    return rows


def save_candidates_batch(
    rows: list[dict[str, str]],
    *,
    filename: str,
) -> dict[str, Any]:
    if not rows:
        raise ValueError("Không có dòng nào để lưu.")

    created = _now()
    candidate_ids: list[int] = []
    with get_conn() as conn:
        cur = conn.execute(
            """
            INSERT INTO import_batches (filename, created_at, status)
            VALUES (?, ?, 'draft')
            """,
            (filename, created),
        )
        batch_id = int(cur.lastrowid)
        for i, row in enumerate(rows, start=1):
            normalized = _payload_for_storage(row)
            status = _submit_status_from_row(row, default="pending")
            cur = conn.execute(
                """
                INSERT INTO candidates (
                    batch_id, row_no, full_name, mobile, email,
                    payload_json, status, error
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    batch_id,
                    i,
                    normalized.get("FullName") or None,
                    normalized.get("Mobile") or None,
                    normalized.get("Email") or None,
                    json.dumps(normalized, ensure_ascii=False),
                    status,
                    (row.get("SubmitError") or row.get("_error") or None) or None,
                ),
            )
            candidate_ids.append(int(cur.lastrowid))
    return {
        "batch_id": batch_id,
        "count": len(rows),
        "filename": filename,
        "candidate_ids": candidate_ids,
    }


def update_candidates_batch(
    batch_id: int,
    rows: list[dict[str, str]],
) -> dict[str, Any]:
    """Upsert candidates in an existing batch — giữ status success/failed theo dòng."""
    if not rows:
        raise ValueError("Không có dòng nào để lưu.")

    with get_conn() as conn:
        exists = conn.execute(
            "SELECT id, filename FROM import_batches WHERE id = ?",
            (batch_id,),
        ).fetchone()
        if not exists:
            raise ValueError(f"Không tìm thấy batch #{batch_id}.")

        existing = conn.execute(
            """
            SELECT id, status, error, response_body, submitted_at
            FROM candidates WHERE batch_id = ?
            """,
            (batch_id,),
        ).fetchall()
        by_id = {int(r["id"]): dict(r) for r in existing}

        kept_ids: list[int] = []
        for i, row in enumerate(rows, start=1):
            normalized = _payload_for_storage(row)
            status = _submit_status_from_row(row, default="pending")
            error = (row.get("SubmitError") or row.get("_error") or "") or None
            cid_raw = row.get("_candidate_id") or row.get("_id") or ""
            try:
                cid = int(str(cid_raw).strip()) if str(cid_raw).strip() else None
            except ValueError:
                cid = None

            if cid is not None and cid in by_id:
                old = by_id[cid]
                # Giữ response/submitted_at nếu vẫn cùng trạng thái DB
                if status == (old.get("status") or ""):
                    response_body = old.get("response_body")
                    submitted_at = old.get("submitted_at")
                    if not error:
                        error = old.get("error")
                else:
                    response_body = None
                    submitted_at = None
                conn.execute(
                    """
                    UPDATE candidates
                    SET row_no = ?, full_name = ?, mobile = ?, email = ?,
                        payload_json = ?, status = ?, error = ?,
                        response_body = ?, submitted_at = ?
                    WHERE id = ? AND batch_id = ?
                    """,
                    (
                        i,
                        normalized.get("FullName") or None,
                        normalized.get("Mobile") or None,
                        normalized.get("Email") or None,
                        json.dumps(normalized, ensure_ascii=False),
                        status,
                        error,
                        response_body,
                        submitted_at,
                        cid,
                        batch_id,
                    ),
                )
                kept_ids.append(cid)
            else:
                cur = conn.execute(
                    """
                    INSERT INTO candidates (
                        batch_id, row_no, full_name, mobile, email,
                        payload_json, status, error
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        batch_id,
                        i,
                        normalized.get("FullName") or None,
                        normalized.get("Mobile") or None,
                        normalized.get("Email") or None,
                        json.dumps(normalized, ensure_ascii=False),
                        status,
                        error,
                    ),
                )
                kept_ids.append(int(cur.lastrowid))

        orphan_ids = [eid for eid in by_id if eid not in set(kept_ids)]
        if orphan_ids:
            placeholders = ",".join("?" * len(orphan_ids))
            conn.execute(
                f"DELETE FROM candidates WHERE batch_id = ? AND id IN ({placeholders})",
                (batch_id, *orphan_ids),
            )

        # Cập nhật trạng thái batch theo kết quả từng dòng
        statuses = [
            r["status"]
            for r in conn.execute(
                "SELECT status FROM candidates WHERE batch_id = ?",
                (batch_id,),
            ).fetchall()
        ]
        batch_status = _batch_status_from_candidate_statuses(statuses)
        conn.execute(
            "UPDATE import_batches SET status = ? WHERE id = ?",
            (batch_status, batch_id),
        )

    return {
        "batch_id": batch_id,
        "count": len(rows),
        "filename": exists["filename"],
        "candidate_ids": kept_ids,
    }


def _payload_for_storage(row: dict[str, Any]) -> dict[str, str]:
    normalized = {
        col: str(row.get(col) or "") for col in PAYLOAD_COLUMNS
    }
    normalized.update(resolve_location_fields(normalized))
    return normalized


def _submit_status_from_row(row: dict[str, Any], *, default: str = "pending") -> str:
    raw = str(row.get("SubmitStatus") or row.get("_status") or default).strip().lower()
    if raw in {"pending", "success", "failed", "invalid"}:
        return raw
    return default


def _batch_status_from_candidate_statuses(statuses: list[str]) -> str:
    if not statuses:
        return "draft"
    normalized = [(s or "pending").lower() for s in statuses]
    if all(s == "pending" for s in normalized):
        return "draft"
    if all(s == "success" for s in normalized):
        return "done"
    if any(s == "failed" for s in normalized):
        return "done_with_errors"
    if any(s == "pending" for s in normalized):
        return "partial"
    return "done"


def load_batch_as_draft(batch_id: int) -> tuple[list[dict[str, str]], str]:
    """Load saved batch → payload rows + filename for the review grid."""
    items = list_imported_candidates(batch_id=batch_id)
    if not items:
        raise ValueError(f"Batch #{batch_id} không có ứng viên.")
    rows = []
    for item in items:
        row = {col: str(item.get(col) or "") for col in PAYLOAD_COLUMNS}
        row["_candidate_id"] = str(item.get("_id") or "")
        row["SubmitStatus"] = str(item.get("_status") or "pending")
        row["SubmitError"] = str(item.get("_error") or "")
        rows.append(enrich_row_address(row, force=False))
    filename = str(items[0].get("_filename") or f"batch_{batch_id}.xlsx")
    return rows, filename


def list_imported_candidates(batch_id: int | None = None) -> list[dict[str, Any]]:
    with get_conn() as conn:
        if batch_id is None:
            rows = conn.execute(
                """
                SELECT c.id, c.batch_id, c.row_no, c.status, c.error,
                       c.response_body, c.submitted_at, c.payload_json,
                       b.filename, b.created_at
                FROM candidates c
                JOIN import_batches b ON b.id = c.batch_id
                ORDER BY c.batch_id DESC, c.row_no ASC
                """
            ).fetchall()
        else:
            rows = conn.execute(
                """
                SELECT c.id, c.batch_id, c.row_no, c.status, c.error,
                       c.response_body, c.submitted_at, c.payload_json,
                       b.filename, b.created_at
                FROM candidates c
                JOIN import_batches b ON b.id = c.batch_id
                WHERE c.batch_id = ?
                ORDER BY c.row_no ASC
                """,
                (batch_id,),
            ).fetchall()

    out: list[dict[str, Any]] = []
    for r in rows:
        payload = json.loads(r["payload_json"] or "{}")
        item = {col: str(payload.get(col) or "") for col in PAYLOAD_COLUMNS}
        item["_id"] = r["id"]
        item["_batch_id"] = r["batch_id"]
        item["_row_no"] = r["row_no"]
        item["_status"] = r["status"]
        item["_error"] = r["error"] or ""
        item["_response_body"] = r["response_body"] or ""
        item["_submitted_at"] = r["submitted_at"] or ""
        item["_filename"] = r["filename"]
        item["_created_at"] = r["created_at"]
        out.append(item)
    return out


def list_import_batches() -> list[dict[str, Any]]:
    with get_conn() as conn:
        rows = conn.execute(
            """
            SELECT b.id, b.filename, b.created_at, b.status,
                   COUNT(c.id) AS candidate_count,
                   SUM(CASE WHEN c.status = 'success' THEN 1 ELSE 0 END) AS success_count,
                   SUM(CASE WHEN c.status = 'failed' THEN 1 ELSE 0 END) AS failed_count,
                   SUM(CASE WHEN c.status = 'pending' OR c.status IS NULL THEN 1 ELSE 0 END)
                     AS pending_count
            FROM import_batches b
            LEFT JOIN candidates c ON c.batch_id = b.id
            GROUP BY b.id
            ORDER BY b.id DESC
            """
        ).fetchall()
    return [dict(r) for r in rows]


def delete_import_batch(batch_id: int) -> dict[str, Any]:
    """Xóa batch và toàn bộ ứng viên thuộc batch."""
    with get_conn() as conn:
        row = conn.execute(
            "SELECT id, filename FROM import_batches WHERE id = ?",
            (batch_id,),
        ).fetchone()
        if row is None:
            raise ValueError(f"Không tìm thấy batch #{batch_id}.")
        count = conn.execute(
            "SELECT COUNT(*) AS n FROM candidates WHERE batch_id = ?",
            (batch_id,),
        ).fetchone()["n"]
        conn.execute("DELETE FROM candidates WHERE batch_id = ?", (batch_id,))
        conn.execute("DELETE FROM import_batches WHERE id = ?", (batch_id,))
    return {
        "batch_id": batch_id,
        "filename": row["filename"],
        "deleted_candidates": int(count),
    }


def export_imported_candidates_bytes(batch_id: int | None = None) -> bytes:
    items = list_imported_candidates(batch_id=batch_id)
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidates"
    for idx, col in enumerate(CANDIDATE_TEMPLATE_COLUMNS, start=1):
        ws.cell(row=1, column=idx, value=col)
    _style_header(ws, 1)

    for r_i, item in enumerate(items, start=2):
        for c_i, col in enumerate(CANDIDATE_TEMPLATE_COLUMNS, start=1):
            ws.cell(row=r_i, column=c_i, value=item.get(col, ""))

    meta = wb.create_sheet("Meta")
    meta["A1"] = "ExportedAt"
    meta["B1"] = _now()
    meta["A2"] = "BatchId"
    meta["B2"] = batch_id if batch_id is not None else "all"
    meta["A3"] = "Count"
    meta["B3"] = len(items)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
